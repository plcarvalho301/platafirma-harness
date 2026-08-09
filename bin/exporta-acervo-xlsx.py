#!/usr/bin/env python3
"""Exporta o schema `acervo` (rag_extractor) para um .xlsx, uma aba por tabela.

SoT da classificacao continua sendo platafirma-conhecimento/ontologia/acervo/*.jsonl.
Este arquivo e uma FOTO do SoR (Postgres) para leitura fora da bancada, nao fonte.
"""
import csv
import io
import subprocess
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CONTAINER = "rag-extractor-pg"
DB = "rag_extractor"
USER = "rag"
SCHEMA = "acervo"
OUT = f"/tmp/acervo-{date.today().isoformat()}.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def psql(sql: str) -> str:
    res = subprocess.run(
        ["docker", "exec", CONTAINER, "psql", "-U", USER, "-d", DB,
         "--csv", "-t", "-A", "-c", sql],
        capture_output=True, text=True, check=True,
    )
    return res.stdout


def psql_csv(sql: str) -> str:
    res = subprocess.run(
        ["docker", "exec", CONTAINER, "psql", "-U", USER, "-d", DB,
         "--csv", "-c", sql],
        capture_output=True, text=True, check=True,
    )
    return res.stdout


def tabelas() -> list[str]:
    sql = (
        "SELECT table_name FROM information_schema.tables "
        f"WHERE table_schema = '{SCHEMA}' AND table_type = 'BASE TABLE' "
        "ORDER BY table_name"
    )
    return [l for l in psql(sql).splitlines() if l.strip()]


def main() -> int:
    wb = Workbook()
    wb.remove(wb.active)
    resumo = []

    for tabela in tabelas():
        dados = psql_csv(f'SELECT * FROM {SCHEMA}."{tabela}"')
        linhas = list(csv.reader(io.StringIO(dados)))
        if not linhas:
            continue

        # nome de aba: 31 chars, sem os caracteres proibidos do Excel
        aba = tabela[:31]
        ws = wb.create_sheet(aba)

        for linha in linhas:
            ws.append(linha)

        # cabecalho
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # largura por coluna, com teto para o texto longo (definicao, resumo)
        for idx, col in enumerate(ws.iter_cols(), start=1):
            largura = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[get_column_letter(idx)].width = min(max(largura + 2, 10), 60)

        resumo.append((tabela, len(linhas) - 1, len(linhas[0])))

    # aba de indice, primeira posicao
    idx_ws = wb.create_sheet("_indice", 0)
    idx_ws.append(["tabela", "linhas", "colunas"])
    for nome, nlin, ncol in resumo:
        idx_ws.append([nome, nlin, ncol])
    idx_ws.append([])
    idx_ws.append(["Foto do schema acervo (Postgres rag_extractor)."])
    idx_ws.append(["SoT: platafirma-conhecimento/ontologia/acervo/*.jsonl (git)."])
    idx_ws.append([f"Extraido em {date.today().isoformat()}."])
    for cell in idx_ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for col, w in (("A", 22), ("B", 10), ("C", 10)):
        idx_ws.column_dimensions[col].width = w

    wb.save(OUT)
    print(OUT)
    for nome, nlin, ncol in resumo:
        print(f"  {nome}: {nlin} linhas, {ncol} colunas")
    return 0


if __name__ == "__main__":
    sys.exit(main())
